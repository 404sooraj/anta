"""
Seed MongoDB from Excel files in server/sample_data/.

Mapping (see sample_data/README.md):
- Partners.xlsx (sheet: result) -> stations: id->station_id, latitude, longitude->location, name
- Call Recording .xlsx (sheet: Sheet1) -> conversations + users: Date->start_time, Calling No.->user_id
- ChargingEvents.xlsx (sheet: result) -> swaps + users: deviceId->user_id, date->date, station_id
- BatteryLogs.xlsx: not mapped to the 8-collection model (optional future use)

Usage:
  uv run python -m scripts.seed_from_excel
  uv run python -m scripts.seed_from_excel --drop   # drop collections before insert
"""

import argparse
import asyncio
import logging
import os
import uuid
from datetime import datetime, timezone
from pathlib import Path

from dotenv import load_dotenv
from motor.motor_asyncio import AsyncIOMotorClient
from openpyxl import load_workbook

load_dotenv()

# Add server root to path so we can import db
SERVER_ROOT = Path(__file__).resolve().parent.parent
os.chdir(SERVER_ROOT)

from db.connection import DB_NAME

logging.basicConfig(level=logging.INFO, format="%(levelname)s: %(message)s")
logger = logging.getLogger(__name__)

SAMPLE_DATA_DIR = SERVER_ROOT / "sample_data"


def _excel_date_to_datetime(val):
    """Convert Excel date number or datetime to datetime."""
    if val is None:
        return None
    if isinstance(val, datetime):
        return val
    if hasattr(val, "isoformat"):
        return val
    try:
        import datetime as dt
        if isinstance(val, (int, float)):
            # Excel serial date: days since 1899-12-30
            return dt.datetime(1899, 12, 30) + dt.timedelta(days=float(val))
    except Exception:
        pass
    return None


async def seed_stations(db, drop: bool) -> None:
    """Partners.xlsx -> stations. Columns: id, latitude, longitude, isActiveDsk."""
    path = SAMPLE_DATA_DIR / "Partners.xlsx"
    if not path.exists():
        logger.warning("Partners.xlsx not found, skipping stations")
        return
    if drop:
        await db.stations.delete_many({})
    wb = load_workbook(path, read_only=True)
    sheet = wb["result"]
    rows = list(sheet.iter_rows(min_row=2, values_only=True))
    headers = ["id", "latitude", "longitude", "isActiveDsk"]
    docs = []
    for row in rows:
        if not row or row[0] is None:
            continue
        row = list(row)[:4]
        d = dict(zip(headers, row))
        sid = str(d.get("id") or "").strip()
        if not sid:
            continue
        try:
            lat = float(d.get("latitude") or 0)
            lon = float(d.get("longitude") or 0)
            location = {"type": "Point", "coordinates": [lon, lat]}
        except (TypeError, ValueError):
            location = {"type": "Point", "coordinates": [0, 0]}
        docs.append({
            "station_id": sid,
            "name": f"Station {sid}",
            "location": location,
            "available_batteries": 0,
        })
    wb.close()
    if docs:
        await db.stations.insert_many(docs)
        logger.info("Inserted %d stations from Partners.xlsx", len(docs))


async def ensure_user(db, user_id: str, name: str = "", phone: str = "") -> None:
    """Insert user if not exists."""
    existing = await db.users.find_one({"user_id": user_id})
    if existing:
        return
    await db.users.insert_one({
        "user_id": user_id,
        "name": name or user_id,
        "phone_number": phone or user_id,
        "language": "en",
    })
    logger.debug("Created user %s", user_id)


async def seed_conversations(db, drop: bool) -> None:
    """Call Recording .xlsx -> conversations. Columns: Date, Name, Issue type, Call recording link, Calling No."""
    path = SAMPLE_DATA_DIR / "Call Recording .xlsx"
    if not path.exists():
        logger.warning("Call Recording .xlsx not found, skipping conversations")
        return
    if drop:
        await db.conversations.delete_many({})
    wb = load_workbook(path, read_only=True)
    sheet = wb["Sheet1"]
    rows = list(sheet.iter_rows(min_row=2, values_only=True))
    headers = ["Date", "Name", "Issue type", "Call recording link", "Calling No."]
    count = 0
    for i, row in enumerate(rows):
        if not row or row[0] is None:
            continue
        row = list(row)[:5]
        d = dict(zip(headers, row))
        calling_no = str(d.get("Calling No.") or d.get("Name") or "").strip() or f"call_{i}"
        await ensure_user(db, calling_no, name=str(d.get("Name") or ""), phone=str(d.get("Calling No.") or ""))
        start_time = _excel_date_to_datetime(d.get("Date"))
        if start_time is None:
            start_time = datetime.now(timezone.utc)
        session_id = str(uuid.uuid4())
        await db.conversations.insert_one({
            "session_id": session_id,
            "user_id": calling_no,
            "language": "en",
            "start_time": start_time,
            "end_time": None,
            "outcome": None,
        })
        count += 1
    wb.close()
    logger.info("Inserted %d conversations from Call Recording .xlsx", count)


async def seed_swaps(db, drop: bool) -> None:
    """ChargingEvents.xlsx -> swaps. Columns: date, deviceId, ts, lat, lon, soc, ..."""
    path = SAMPLE_DATA_DIR / "ChargingEvents.xlsx"
    if not path.exists():
        logger.warning("ChargingEvents.xlsx not found, skipping swaps")
        return
    if drop:
        await db.swaps.delete_many({})
    # Get first station for FK and station_snapshot (denormalized for single-doc read)
    first_station = await db.stations.find_one({})
    station_id = first_station["station_id"] if first_station else "unknown"
    station_snapshot = (
        {"name": first_station["name"], "location": first_station["location"]}
        if first_station
        else None
    )
    wb = load_workbook(path, read_only=True)
    sheet = wb["result"]
    rows = list(sheet.iter_rows(min_row=2, values_only=True))
    headers = ["date", "deviceId", "ts", "lat", "lon", "soc", "discharging_time", "charge_start_time"]
    count = 0
    for row in rows:
        if not row or row[0] is None:
            continue
        row = list(row)[:8]
        d = dict(zip(headers, row))
        device_id = str(d.get("deviceId") or "").strip()
        if not device_id:
            continue
        await ensure_user(db, device_id)
        dt_val = _excel_date_to_datetime(d.get("date")) or _excel_date_to_datetime(d.get("ts"))
        if dt_val is None:
            dt_val = datetime.now(timezone.utc)
        swap_id = str(uuid.uuid4())
        swap_doc = {
            "swap_id": swap_id,
            "user_id": device_id,
            "station_id": station_id,
            "date": dt_val,
            "amount": 1,
        }
        if station_snapshot is not None:
            swap_doc["station_snapshot"] = station_snapshot
        await db.swaps.insert_one(swap_doc)
        count += 1
    wb.close()
    logger.info("Inserted %d swaps from ChargingEvents.xlsx", count)


async def seed_agents(db, drop: bool) -> None:
    """Seed placeholder agents (no Excel source)."""
    if drop:
        await db.agents.delete_many({})
    agents = [
        {"agent_id": "agent_1", "name": "Support Agent 1"},
        {"agent_id": "agent_2", "name": "Support Agent 2"},
    ]
    for a in agents:
        existing = await db.agents.find_one({"agent_id": a["agent_id"]})
        if not existing:
            await db.agents.insert_one(a)
    logger.info("Seeded %d agents", len(agents))


async def run(drop: bool) -> None:
    url = os.getenv("MONGODB_URL", "mongodb://localhost:27017")
    client = AsyncIOMotorClient(url)
    db = client[DB_NAME]
    try:
        if drop:
            for name in ["users", "stations", "conversations", "agents", "swaps", "intent_logs", "handoffs", "subscriptions"]:
                try:
                    await db.drop_collection(name)
                    logger.info("Dropped collection %s", name)
                except Exception:
                    pass  # collection may not exist
        await seed_stations(db, drop)
        await seed_agents(db, drop)
        await seed_conversations(db, drop)
        await seed_swaps(db, drop)
    finally:
        client.close()
    logger.info("Seed complete.")


def main() -> None:
    parser = argparse.ArgumentParser(description="Seed MongoDB from sample_data Excel files")
    parser.add_argument("--drop", action="store_true", help="Drop collections before inserting")
    args = parser.parse_args()
    asyncio.run(run(drop=args.drop))


if __name__ == "__main__":
    main()
